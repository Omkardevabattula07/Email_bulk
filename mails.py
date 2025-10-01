import smtplib
from openpyxl import load_workbook
import tkinter as tk
from tkinter import messagebox, filedialog
from tkinter import ttk
import os
from dotenv import load_dotenv

# Load env
load_dotenv()

SENDER_EMAIL = os.getenv("SENDER_EMAIL")
APP_PASSWORD = os.getenv("APP_PASSWORD")

email_list_path = None
attachments = []

def select_email_list():
    global email_list_path
    email_list_path = filedialog.askopenfilename(
        title="Select Email List",
        filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
    )
    if email_list_path:
        email_list_label.config(text=os.path.basename(email_list_path))
    else:
        email_list_label.config(text="No file selected")

def select_attachments():
    global attachments
    files = filedialog.askopenfilenames(
        title="Select Attachments"
    )
    attachments = list(files)
    if attachments:
        attachments_label.config(text=f"{len(attachments)} file(s) selected")
    else:
        attachments_label.config(text="No files selected")

def clear_fields():
    subject_entry.delete(0, tk.END)
    text_box.delete("1.0", tk.END)
    global email_list_path, attachments
    email_list_path = None
    attachments = []
    email_list_label.config(text="No file selected")
    attachments_label.config(text="No files selected")

def send_emails():
    subject = subject_entry.get().strip()
    message = text_box.get("1.0", tk.END).strip()
    if not subject:
        messagebox.showerror("Error", "Subject is empty!")
        return
    if not message:
        messagebox.showerror("Error", "Message box is empty!")
        return
    if not email_list_path:
        messagebox.showerror("Error", "No email list selected!")
        return

    try:
        emails = []
        if email_list_path.endswith(".xlsx"):
            wb = load_workbook(email_list_path)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                email = row[0]
                if email:
                    emails.append(email)
        elif email_list_path.endswith(".csv"):
            import csv
            with open(email_list_path, newline='') as csvfile:
                reader = csv.reader(csvfile)
                next(reader)  # skip header
                for row in reader:
                    if row and row[0]:
                        emails.append(row[0])

        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(SENDER_EMAIL, APP_PASSWORD)

        for email in emails:
            from email.message import EmailMessage
            msg = EmailMessage()
            msg["From"] = SENDER_EMAIL
            msg["To"] = email
            msg["Subject"] = subject
            msg.set_content(message)
            for file_path in attachments:
                with open(file_path, "rb") as f:
                    file_data = f.read()
                    file_name = os.path.basename(file_path)
                    msg.add_attachment(file_data, maintype="application", subtype="octet-stream", filename=file_name)
            server.send_message(msg)

        server.quit()
        messagebox.showinfo("Success", "Emails sent successfully!")

    except Exception as e:
        messagebox.showerror("Error", str(e))

# --- UI ---
root = tk.Tk()
root.title("Bulk Email Sender")

# Set window size and center it with space around
window_width = 500
window_height = 420
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
x = int((screen_width / 2) - (window_width / 2))
y = int((screen_height / 2) - (window_height / 2))
root.geometry(f"{window_width}x{window_height}+{x}+{y}")
root.resizable(False, False)

main_frame = ttk.Frame(root, padding="20 20 20 20")
main_frame.pack(fill=tk.BOTH, expand=True)

ttk.Label(main_frame, text="Bulk Email Sender", font=("Segoe UI", 18, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 15))

ttk.Label(main_frame, text="Subject:").grid(row=1, column=0, sticky="w")
subject_entry = ttk.Entry(main_frame, width=45)
subject_entry.grid(row=1, column=1, pady=5, sticky="ew")

ttk.Label(main_frame, text="Message:").grid(row=2, column=0, sticky="nw")
text_box = tk.Text(main_frame, height=7, width=45, font=("Segoe UI", 10))
text_box.grid(row=2, column=1, pady=5, sticky="ew")

ttk.Button(main_frame, text="Select Email List", command=select_email_list).grid(row=3, column=0, pady=5, sticky="w")
email_list_label = ttk.Label(main_frame, text="No file selected", foreground="gray")
email_list_label.grid(row=3, column=1, sticky="w")

ttk.Button(main_frame, text="Select Attachments", command=select_attachments).grid(row=4, column=0, pady=5, sticky="w")
attachments_label = ttk.Label(main_frame, text="No files selected", foreground="gray")
attachments_label.grid(row=4, column=1, sticky="w")

button_frame = ttk.Frame(main_frame)
button_frame.grid(row=5, column=0, columnspan=2, pady=20)

send_button = ttk.Button(button_frame, text="Send Emails", command=send_emails)
send_button.grid(row=0, column=0, padx=10)

clear_button = ttk.Button(button_frame, text="Clear Fields", command=clear_fields)
clear_button.grid(row=0, column=1, padx=10)

for i in range(6):
    main_frame.rowconfigure(i, weight=1)
main_frame.columnconfigure(1, weight=1)

root.mainloop()
